import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import sys
import os
import numpy

# Initializing empty lists

page_link = []
articles_list = []
link_list = []


def integer_input_validator(input_variable):

    # Function to check if input is an integer or not

    try:
        int(input_variable)
    except ValueError:
        print("Invalid input! Please enter integer type values.")
        sys.exit(0)


# Input to fetch true or false news

news_type = input(
    "Select if you want to fetch for true news or false/parody/satire news. (Enter 1 for true snd 0 false): ")

integer_input_validator(news_type)

news_type = int(news_type)

if news_type == 0:
    file_name = 'False News.xlsx'

    # Input to decide which website to fetch from

    website = input(
        "Choose website to fetch news from. (Enter 1 for The Fauxy, enter 2 for The Onion): ")
    integer_input_validator(website)

    # Input the number of pages to fetch

    number_of_pages = input(
        "Enter number of pages to fetch articles from (Max limit is 40 for the Fauxy, 118 for The Onion): ")
    integer_input_validator(number_of_pages)


elif news_type == 1:
    file_name = 'True News.xlsx'

    # Input to decide which website to fetch from

    website = input(
        "Choose website to fetch news from. (Enter 1 for Hindustan Times, enter 2 for NDTV or enter 3 for The Indian Express): ")
    integer_input_validator(website)

    # Input the number of pages to fetch

    number_of_pages = input(
        "Enter number of pages to fetch articles from (Max limit is 50 for Hindustan Times, 14 for NDTV and 100 for The Indian Express): ")
    integer_input_validator(number_of_pages)

else:
    print("Invalid input! Please input either 1 or 0.")
    sys.exit(0)

website = int(website)

if website == 1 and news_type == 1:
    website = "hindustantimes"
elif website == 2 and news_type == 1:
    website = "ndtv"
elif website == 3 and news_type == 1:
    website = "theindianexpress"
elif website == 1 and news_type == 0:
    website = "thefauxy"
elif website == 2 and news_type == 0:
    website = "theonion"
else:
    print("Invalid input! Please enter values from the given options.")
    sys.exit(0)

# Input to decide the topic of news

category = input(
    f'Please enter the topic keyword corresponding to the website {website} for fetching articles: ')

number_of_pages = int(number_of_pages)

if website == "hindustantimes" and number_of_pages > 50:
    print("Number of pages are out of range. Value set to maximum (50)")
    number_of_pages = 50
elif website == "ndtv" and number_of_pages > 14:
    print("Number of pages are out of range. Value set to maximum (14)")
    number_of_pages = 14
elif website == "theindianexpress" and number_of_pages > 100:
    print("Number of pages are out of range. Value set to maximum (100)")
    number_of_pages = 100
elif website == "thefauxy" and number_of_pages > 40:
    print("Number of pages are out of range. Value set to maximum (40)")
    number_of_pages = 40
elif website == "theonion" and number_of_pages > 118:
    print("Number of pages are out of range. Value set to maximum (118)")
    number_of_pages = 118

# Declaring headers to be used when visiting websites to avoid bot detection

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36",
           "X-Amzn-Trace-Id": "Root=1-62d8036d-2b173c1f2e4e7a416cc9e554", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
           "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB", }

# Dictionary containing values corresponding to different websites for scraping data

fetch_values_dictionary = {
    "hindustantimes": {
        "domain_link": "https://www.hindustantimes.com/",
        "link": f'https://www.hindustantimes.com/{category}/page-',
        "range_start": 1,
        "range_end": 1 + number_of_pages,
        "range_offset": 1,
        "articles_find_element": {
            "element": "section",
            "element_class": "listingPage"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "cartHolder",
        },
        "link_find_element": {
            "element": "h3",
            "element_class": "hdg3",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "detailPage",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "hdg1",
        },
        "datetime_find_element": {
            "element": "div",
            "element_class": "dateTime",
        },
    },
    "ndtv": {
        "domain_link": "https://www.ndtv.com/",
        "link": f'https://www.ndtv.com/{category}/page-',
        "range_start": 1,
        "range_end": 1 + number_of_pages,
        "range_offset": 1,
        "articles_find_element": {
            "element": "div",
            "element_class": "lisingNews"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "news_Itm",
        },
        "link_find_element": {
            "element": "h2",
            "element_class": "newsHdng",
        },
        "content_find_element": {
            "element": "section",
            "element_class": "col-900",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "sp-ttl",
        },
        "datetime_find_element": {
            "element": "span",
            "element_itemprop": "dateModified",
        },
    },
    "theindianexpress": {
        "domain_link": "https://indianexpress.com/",
        "link": f'https://indianexpress.com/section/{category}/page/',
        "range_start": 2,
        "range_end": 2 + number_of_pages,
        "range_offset": 1,
        "articles_find_element": {
            "element": "div",
            "element_class": "nation"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "articles",
        },
        "link_find_element": {
            "element": "div",
            "element_class": "snaps",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "container native_story",
        },
        "title_find_element": {
            "element": "h1",
            "element_itemprop": "headline",
        },
        "datetime_find_element": {
            "element": "span",
            "element_itemprop": "dateModified",
        },
    },
    "thefauxy": {
        "domain_link": "https://thefauxy.com/",
        "link": f'https://thefauxy.com/{category}/page/',
        "range_start": 1,
        "range_end": 1 + number_of_pages,
        "range_offset": 1,
        "articles_find_element": {
            "element": "div",
            "element_class": "entries"
        },
        "articles_findall_element": {
            "element": "article",
            "element_class": "entry-card",
        },
        "link_find_element": {
            "element": "h2",
            "element_class": "entry-title",
        },
        "content_find_element": {
            "element": "article",
            "element_class": "post",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "page-title",
        },
        "datetime_find_element": {
            "element": "time",
            "element_class": "ct-meta-element-date",
        },
    },
    "theonion": {
        "domain_link": "https://www.theonion.com/",
        "link": f'https://www.theonion.com/breaking-news/{category}?startIndex=',
        "range_start": 00,
        "range_end": 00 + number_of_pages*20,
        "range_offset": 20,
        "articles_find_element": {
            "element": "div",
            "element_class": "sc-17uq8ex-0"
        },
        "articles_findall_element": {
            "element": "article",
            "element_class": "js_post_item",
        },
        "link_find_element": {
            "element": "div",
            "element_class": "sc-cw4lnv-5",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "sc-101yw2y-9",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "sc-1efpnfq-0",
        },
        "datetime_find_element": {
            "element": "time",
            "element_class": "sc-uhd9ir-0",
        },
    },
}

# If excel file already exists, activate it and append new data in it

if os.path.exists(file_name):
    print(f'File {file_name} detetcted. Appending data to {file_name}')
    excel = openpyxl.load_workbook(file_name)
    sheet = excel.active

# Else create a new file

else:
    print(f'Creating new file named {file_name}')
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = 'News'
    sheet.append(['Title', 'Text', 'Subject', 'Date'])


def fetch_news(fetch_values_dictionary, category, website):

    # Funtion to fetch data from websites

    # Editing category variable to be appended in the sheet

    if len(re.findall('news', category)) == 0:
        category = f'{category} News'

    category = category.replace('-', ' ')
    category = category.title()

    def article_fetcher():

        # Function to fetch and create a list of news articles

        print("Fetching Articles...")

        for i in range(fetch_values_dictionary[website].get("range_start"),
                       fetch_values_dictionary[website].get("range_end"), fetch_values_dictionary[website].get("range_offset")):
            page_link = (
                f'{fetch_values_dictionary[website].get("link")}{i}')

            i = i+1
            try:
                page = requests.get(page_link, headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                articles = soup2.find(fetch_values_dictionary[website]["articles_find_element"].get("element"),
                                      class_=fetch_values_dictionary[website]["articles_find_element"].get("element_class")).find_all(
                    fetch_values_dictionary[website]["articles_findall_element"].get(
                        "element"),
                    class_=fetch_values_dictionary[website]["articles_findall_element"].get("element_class"))

                articles_list.extend(articles)

            except Exception as e:
                print(e)

    article_fetcher()

    def link_list_maker():

        # Function to extract links of all the articles from article list

        print("Getting Links...")

        for article in articles_list:

            link = article.find(fetch_values_dictionary[website]["link_find_element"].get("element"),
                                class_=fetch_values_dictionary[website]["link_find_element"].get("element_class"))

            if link is not None:
                link = link.find('a').get('href')

                if (len(re.findall(fetch_values_dictionary[website].get("domain_link"), link)) == 0):
                    link = f'{fetch_values_dictionary[website].get("domain_link")}{link}'
                link_list.append(link)

    link_list_maker()

    def content_fetcher():

        # Function to extract news data from each article

        print("Fetching Content...")

        for i in range(len(link_list)):
            try:
                page = requests.get(link_list[i], headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                if fetch_values_dictionary[website]["content_find_element"].get("element_class") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        class_=fetch_values_dictionary[website]["content_find_element"].get("element_class"))
                elif fetch_values_dictionary[website]["content_find_element"].get("element_id") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        id=fetch_values_dictionary[website]["content_find_element"].get("element_id"))
                elif fetch_values_dictionary[website]["content_find_element"].get("element_itemprop") is not None:
                    content = soup2.find(
                        fetch_values_dictionary[website]["content_find_element"].get(
                            "element"),
                        itemprop=fetch_values_dictionary[website]["content_find_element"].get("element_itemprop"))

                try:
                    if fetch_values_dictionary[website]["title_find_element"].get("element_class") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            class_=fetch_values_dictionary[website]["title_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary[website]["title_find_element"].get("element_id") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            id=fetch_values_dictionary[website]["title_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary[website]["title_find_element"].get("element_itemprop") is not None:
                        title = content.find(
                            fetch_values_dictionary[website]["title_find_element"].get(
                                "element"),
                            itemprop=fetch_values_dictionary[website]["title_find_element"].get("element_itemprop")).get_text().strip()

                    title = ' '.join(title.split())

                except AttributeError:
                    title = "Null"

                try:
                    if fetch_values_dictionary[website]["datetime_find_element"].get("element_class") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), class_=fetch_values_dictionary[website]["datetime_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary[website]["datetime_find_element"].get("element_id") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), id=fetch_values_dictionary[website]["datetime_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary[website]["datetime_find_element"].get("element_itemprop") is not None:
                        date_time = content.find(
                            fetch_values_dictionary[website]["datetime_find_element"].get(
                                "element"), itemprop=fetch_values_dictionary[website]["datetime_find_element"].get("element_itemprop")).get_text().strip()

                    date_time = date_time.replace('Updated: ', '')
                    date_time = date_time.replace('Published: ', '')
                    date_time = date_time.replace('Published ', '')
                    date_time = date_time.replace('Updated ', '')
                    date_time = date_time.replace('Published:', '')
                    date_time = date_time.replace('Updated:', '')
                    date_time = date_time.replace('Published', '')
                    date_time = date_time.replace('Updated', '')
                    date_time = date_time.replace('On ', '')
                    date_time = date_time.replace('On: ', '')
                    date_time = date_time.replace('On:', '')
                    date_time = date_time.replace('On', '')

                except AttributeError:
                    date_time = "Null"

                try:
                    body = [x.get_text() for x in content.find_all('p')]
                    body = ' '.join(body)
                    body = ' '.join(body.split())
                    body = body.replace(' Watch Live News: Follow Us:', '')
                    body = body.replace(' ...view detail', '')
                except AttributeError:
                    body = "Null"

                if (title and body and date_time) != "Null":
                    sheet.append([title, body, category, date_time])

                # print(title, date_time, category, body)

            except Exception as e:
                print(e)

    content_fetcher()


fetch_news(fetch_values_dictionary, category, website)

# Saving changes to the excel file

print("Saving Excel File...")
excel.save(file_name)
